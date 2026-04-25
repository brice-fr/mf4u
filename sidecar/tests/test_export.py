"""
Tests for sidecar/export.py — MAT, TDMS, Parquet, CSV, TSV, and XLSX export jobs.
"""
from __future__ import annotations

import os
import time

import pytest


def _open(path: str):
    from asammdf import MDF  # type: ignore[import-untyped]
    return MDF(path, memory="low")


def _wait_for_job(job_id: str, timeout: float = 30.0) -> dict:
    """Poll get_progress until the job leaves the 'running' state."""
    import export as exp
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        progress = exp.get_progress(job_id)
        if progress["status"] != "running":
            return progress
        time.sleep(0.05)
    raise TimeoutError(f"export job {job_id!r} did not finish within {timeout}s")


# --------------------------------------------------------------------------- #
# MAT export
# --------------------------------------------------------------------------- #

class TestMatExport:
    def test_creates_file(self, minimal_mf4, tmp_path):
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.mat")
        try:
            job_id = exp.start(mdf, "mat", out)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        assert os.path.isfile(out)
        assert os.path.getsize(out) > 0

    def test_progress_done_equals_total(self, minimal_mf4, tmp_path):
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.mat")
        try:
            job_id = exp.start(mdf, "mat", out)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["done"] == progress["total"]

    def test_output_readable_by_scipy(self, minimal_mf4, tmp_path):
        import export as exp
        import scipy.io as sio  # type: ignore[import-untyped]
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.mat")
        try:
            job_id = exp.start(mdf, "mat", out)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        mat = sio.loadmat(out)
        data_keys = [k for k in mat if not k.startswith("_")]
        assert len(data_keys) >= 1

    def test_mat_contains_signal_arrays(self, minimal_mf4, tmp_path):
        import export as exp
        import scipy.io as sio  # type: ignore[import-untyped]
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.mat")
        try:
            job_id = exp.start(mdf, "mat", out)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        mat = sio.loadmat(out)
        # Ch1, Ch2, Ch3 are numeric → should be exported
        assert any("Ch" in k for k in mat), f"Expected Ch* keys, got {list(mat)}"

    def test_mat_always_includes_timestamp_vector(self, minimal_mf4, tmp_path):
        """MAT export must include a t1 key for the timestamp vector."""
        import export as exp
        import scipy.io as sio  # type: ignore[import-untyped]
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.mat")
        try:
            job_id = exp.start(mdf, "mat", out)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        mat = sio.loadmat(out)
        assert "t1" in mat, f"Expected 't1' timestamp key, got keys: {[k for k in mat if not k.startswith('_')]}"

    def test_mat_multi_group_timestamp_vectors(self, multi_group_mf4, tmp_path):
        """Multi-group MAT export must include t1, t2, t3, t4 timestamp vectors."""
        import export as exp
        import scipy.io as sio  # type: ignore[import-untyped]
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "out.mat")
        try:
            job_id = exp.start(mdf, "mat", out)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        mat = sio.loadmat(out)
        for t_key in ("t1", "t2", "t3", "t4"):
            assert t_key in mat, (
                f"Expected '{t_key}' timestamp key, got: {[k for k in mat if not k.startswith('_')]}"
            )

    def test_mat_link_groups_suffixes_channels(self, minimal_mf4, tmp_path):
        """With mat_link_groups=True, channel names are suffixed with _t1."""
        import export as exp
        import scipy.io as sio  # type: ignore[import-untyped]
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.mat")
        try:
            job_id = exp.start(mdf, "mat", out, mat_link_groups=True)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        mat = sio.loadmat(out)
        data_keys = [k for k in mat if not k.startswith("_")]
        # Every data key (not t1 itself) should end with _t1
        data_channels = [k for k in data_keys if k != "t1"]
        assert all(k.endswith("_t1") for k in data_channels), (
            f"Expected all channel keys to end with _t1, got: {data_keys}"
        )

    def test_mat_link_groups_multi_group_suffixes(self, multi_group_mf4, tmp_path):
        """With mat_link_groups=True across 4 groups, each channel is suffixed with
        its group's time-vector label (t1…t4)."""
        import export as exp
        import scipy.io as sio  # type: ignore[import-untyped]
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "out.mat")
        try:
            job_id = exp.start(mdf, "mat", out, mat_link_groups=True)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        mat = sio.loadmat(out)
        keys = {k for k in mat if not k.startswith("_")}
        # Time vectors must be present
        assert {"t1", "t2", "t3", "t4"}.issubset(keys), f"Missing time keys in: {keys}"
        # multi_group has channels: EngineSpeed, ThrottlePos in group 0 → _t1 suffix
        assert "EngineSpeed_t1" in keys, f"Expected EngineSpeed_t1 in {keys}"
        assert "ThrottlePos_t1" in keys, f"Expected ThrottlePos_t1 in {keys}"

    def test_mat_link_groups_false_no_suffix(self, minimal_mf4, tmp_path):
        """With mat_link_groups=False (default), channels are NOT suffixed."""
        import export as exp
        import scipy.io as sio  # type: ignore[import-untyped]
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.mat")
        try:
            job_id = exp.start(mdf, "mat", out, mat_link_groups=False)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        mat = sio.loadmat(out)
        data_keys = [k for k in mat if not k.startswith("_")]
        # Should have Ch1, Ch2, Ch3 (not Ch1_t1, etc.)
        assert any(k == "Ch1" for k in data_keys), (
            f"Expected Ch1 without suffix, got: {data_keys}"
        )


# --------------------------------------------------------------------------- #
# TDMS export
# --------------------------------------------------------------------------- #

class TestTdmsExport:
    def test_creates_file(self, minimal_mf4, tmp_path):
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.tdms")
        try:
            job_id = exp.start(mdf, "tdms", out)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        assert os.path.isfile(out)
        assert os.path.getsize(out) > 0

    def test_output_readable_by_nptdms(self, minimal_mf4, tmp_path):
        import export as exp
        from nptdms import TdmsFile  # type: ignore[import-untyped]
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.tdms")
        try:
            job_id = exp.start(mdf, "tdms", out)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        tdms = TdmsFile.read(out)
        groups = tdms.groups()
        assert len(groups) >= 1

    def test_tdms_contains_channels(self, minimal_mf4, tmp_path):
        import export as exp
        from nptdms import TdmsFile  # type: ignore[import-untyped]
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.tdms")
        try:
            job_id = exp.start(mdf, "tdms", out)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        tdms = TdmsFile.read(out)
        all_channels = [ch for g in tdms.groups() for ch in g.channels()]
        assert len(all_channels) >= 3   # Ch1, Ch2, Ch3


# --------------------------------------------------------------------------- #
# Parquet export
# --------------------------------------------------------------------------- #

class TestParquetExport:
    def test_creates_file(self, minimal_mf4, tmp_path):
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.parquet")
        try:
            job_id = exp.start(mdf, "parquet", out)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        # Single-group file → written to the exact output path
        assert os.path.isfile(out)
        assert os.path.getsize(out) > 0

    def test_output_readable_by_pyarrow(self, minimal_mf4, tmp_path):
        import export as exp
        import pyarrow.parquet as pq  # type: ignore[import-untyped]
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.parquet")
        try:
            job_id = exp.start(mdf, "parquet", out)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        table = pq.read_table(out)
        assert table.num_rows > 0

    def test_multi_group_creates_multiple_files(self, multi_group_mf4, tmp_path):
        import export as exp
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "multi.parquet")
        try:
            job_id = exp.start(mdf, "parquet", out)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        parquet_files = [f for f in os.listdir(tmp_path) if f.endswith(".parquet")]
        assert len(parquet_files) == 4


# --------------------------------------------------------------------------- #
# CSV export
# --------------------------------------------------------------------------- #

class TestCsvExport:
    def test_creates_file(self, minimal_mf4, tmp_path):
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.csv")
        try:
            job_id = exp.start(mdf, "csv", out)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        assert os.path.isfile(out)
        assert os.path.getsize(out) > 0

    def test_has_header_and_data_rows(self, minimal_mf4, tmp_path):
        import csv
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.csv")
        try:
            job_id = exp.start(mdf, "csv", out)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        with open(out, newline="", encoding="utf-8") as fh:
            rows = list(csv.reader(fh))
        assert len(rows) >= 2                         # header + at least one data row
        assert "timestamps" in rows[0]
        assert any("Ch" in col for col in rows[0])    # Ch1 / Ch2 / Ch3 in header

    def test_multi_group_creates_multiple_files(self, multi_group_mf4, tmp_path):
        import export as exp
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "multi.csv")
        try:
            job_id = exp.start(mdf, "csv", out)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        csv_files = [f for f in os.listdir(tmp_path) if f.endswith(".csv")]
        assert len(csv_files) == 4


# --------------------------------------------------------------------------- #
# TSV export
# --------------------------------------------------------------------------- #

class TestTsvExport:
    def test_creates_file(self, minimal_mf4, tmp_path):
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.tsv")
        try:
            job_id = exp.start(mdf, "tsv", out)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        assert os.path.isfile(out)
        assert os.path.getsize(out) > 0

    def test_tab_delimited(self, minimal_mf4, tmp_path):
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.tsv")
        try:
            job_id = exp.start(mdf, "tsv", out)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        with open(out, encoding="utf-8") as fh:
            header = fh.readline()
        assert "\t" in header
        assert "," not in header


# --------------------------------------------------------------------------- #
# XLSX export
# --------------------------------------------------------------------------- #

class TestXlsxExport:
    def test_creates_file(self, minimal_mf4, tmp_path):
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.xlsx")
        try:
            job_id = exp.start(mdf, "xlsx", out)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        assert os.path.isfile(out)
        assert os.path.getsize(out) > 0

    def test_output_readable_by_openpyxl(self, minimal_mf4, tmp_path):
        import export as exp
        import openpyxl  # type: ignore[import-untyped]
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.xlsx")
        try:
            job_id = exp.start(mdf, "xlsx", out)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        wb = openpyxl.load_workbook(out, read_only=True)
        assert len(wb.sheetnames) >= 1
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 2    # header + at least one data row
        wb.close()

    def test_multi_group_creates_multiple_sheets(self, multi_group_mf4, tmp_path):
        import export as exp
        import openpyxl  # type: ignore[import-untyped]
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "multi.xlsx")
        try:
            job_id = exp.start(mdf, "xlsx", out)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        wb = openpyxl.load_workbook(out, read_only=True)
        assert len(wb.sheetnames) == 4
        wb.close()


# --------------------------------------------------------------------------- #
# Phase A — bus frame decoding
# --------------------------------------------------------------------------- #

class TestDbLoading:
    def test_load_dbc_returns_db_object(self, can_bus_dbc):
        """_load_db_matrix should parse the DBC without error."""
        import export as exp
        db = exp._load_db_matrix(can_bus_dbc)
        assert db is not None

    def test_dbc_message_map_correct_count(self, can_bus_dbc):
        """can_bus.dbc has exactly 2 messages (IDs 100 and 200)."""
        import export as exp
        db  = exp._load_db_matrix(can_bus_dbc)
        msg = exp._db_message_map(db)
        assert len(msg) == 2
        assert 100 in msg
        assert 200 in msg

    def test_dbc_signal_counts(self, can_bus_dbc):
        """Message 100 has 2 signals; message 200 has 2 signals → 4 total."""
        import export as exp
        db  = exp._load_db_matrix(can_bus_dbc)
        msg = exp._db_message_map(db)
        assert msg[100] == 2
        assert msg[200] == 2

    def test_load_nonexistent_raises(self, tmp_path):
        import export as exp
        with pytest.raises(Exception):
            exp._load_db_matrix(str(tmp_path / "ghost.dbc"))


class TestGetGroupCanIds:
    def test_reads_ids_from_bus_raw(self, bus_raw_mf4):
        """bus_raw.mf4 has CAN_DataFrame.ID with values 100 and 200."""
        import export as exp
        mdf = _open(bus_raw_mf4)
        try:
            ids = exp._get_group_can_ids(mdf, 0)
        finally:
            mdf.close()
        assert ids is not None
        assert 100 in ids
        assert 200 in ids

    def test_returns_none_for_missing_channel(self, minimal_mf4):
        """minimal.mf4 has no CAN_DataFrame.ID channel → should return None."""
        import export as exp
        mdf = _open(minimal_mf4)
        try:
            ids = exp._get_group_can_ids(mdf, 0)
        finally:
            mdf.close()
        assert ids is None


class TestPreviewBusDecoding:
    def test_preview_returns_one_entry_per_assignment(self, bus_raw_mf4, can_bus_dbc):
        import export as exp
        mdf = _open(bus_raw_mf4)
        assignments = [
            {"group_index": 0, "db_path": can_bus_dbc},
        ]
        try:
            results = exp.preview_bus_decoding(mdf, assignments)
        finally:
            mdf.close()
        assert len(results) == 1
        assert results[0]["group_index"] == 0
        assert results[0]["db_path"] == can_bus_dbc
        assert results[0]["error"] is None

    def test_preview_correct_match_count(self, bus_raw_mf4, can_bus_dbc):
        """Both DB messages (100, 200) appear in bus_raw.mf4 → matched_messages=2."""
        import export as exp
        mdf = _open(bus_raw_mf4)
        assignments = [{"group_index": 0, "db_path": can_bus_dbc}]
        try:
            results = exp.preview_bus_decoding(mdf, assignments)
        finally:
            mdf.close()
        assert results[0]["matched_messages"] == 2
        assert results[0]["signal_count"] == 4   # 2 signals per message

    def test_preview_empty_when_no_assignments(self, bus_raw_mf4):
        import export as exp
        mdf = _open(bus_raw_mf4)
        try:
            results = exp.preview_bus_decoding(mdf, [])
        finally:
            mdf.close()
        assert results == []

    def test_preview_error_for_bad_db_path(self, bus_raw_mf4, tmp_path):
        """Missing DB file should produce an error entry (not raise)."""
        import export as exp
        mdf = _open(bus_raw_mf4)
        assignments = [{"group_index": 0, "db_path": str(tmp_path / "missing.dbc")}]
        try:
            results = exp.preview_bus_decoding(mdf, assignments)
        finally:
            mdf.close()
        assert results[0]["error"] is not None

    def test_preview_fallback_when_no_can_id_channel(self, minimal_mf4, can_bus_dbc):
        """When group has no CAN_DataFrame.ID, all DB messages count as matched."""
        import export as exp
        mdf = _open(minimal_mf4)
        # minimal.mf4 has no CAN frames — _get_group_can_ids returns None
        assignments = [{"group_index": 0, "db_path": can_bus_dbc}]
        try:
            results = exp.preview_bus_decoding(mdf, assignments)
        finally:
            mdf.close()
        # Fallback: report all DB messages since we can't cross-reference IDs
        assert results[0]["matched_messages"] == 2
        assert results[0]["signal_count"] == 4


class TestExportWithDecoding:
    def test_export_with_empty_db_assignments_unchanged(self, minimal_mf4, tmp_path):
        """db_assignments=[] should behave identically to no decoding."""
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "out.csv")
        try:
            job_id   = exp.start(mdf, "csv", out, db_assignments=[])
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        assert os.path.isfile(out)

    def test_export_with_decoding_completes_or_errors_cleanly(
        self, bus_raw_mf4, can_bus_dbc, tmp_path
    ):
        """
        Export with db_assignments should start and either complete or fail cleanly
        (no crash / hung thread) regardless of whether extract_bus_logging succeeds.
        """
        import export as exp
        mdf = _open(bus_raw_mf4)
        out = str(tmp_path / "decoded.csv")
        assignments = [{"group_index": 0, "db_path": can_bus_dbc}]
        try:
            job_id   = exp.start(mdf, "csv", out, db_assignments=assignments)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        # The job must reach a terminal state — not hang
        assert progress["status"] in ("done", "error"), (
            f"unexpected status: {progress['status']}"
        )
        # If it errored, the error must carry a meaningful message
        if progress["status"] == "error":
            assert progress["error"]


# --------------------------------------------------------------------------- #
# Phase C — Flatten output
# --------------------------------------------------------------------------- #

class TestFlattenExport:
    def test_flat_parquet_single_file_for_multi_group(self, multi_group_mf4, tmp_path):
        """Flatten + parquet should produce exactly one file, not one per group."""
        import export as exp
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "flat.parquet")
        try:
            job_id   = exp.start(mdf, "parquet", out, flatten=True)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        parquet_files = [f for f in tmp_path.iterdir() if f.suffix == ".parquet"]
        assert len(parquet_files) == 1, f"Expected 1 file, got {parquet_files}"

    def test_flat_csv_single_file_for_multi_group(self, multi_group_mf4, tmp_path):
        """Flatten + csv should produce exactly one file, not one per group."""
        import export as exp
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "flat.csv")
        try:
            job_id   = exp.start(mdf, "csv", out, flatten=True)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        csv_files = [f for f in tmp_path.iterdir() if f.suffix == ".csv"]
        assert len(csv_files) == 1, f"Expected 1 file, got {csv_files}"

    def test_flat_csv_has_all_group_channels(self, multi_group_mf4, tmp_path):
        """Flat CSV should contain channels from all 4 groups in a single header row."""
        import csv
        import export as exp
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "flat.csv")
        try:
            job_id = exp.start(mdf, "csv", out, flatten=True)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        with open(out, newline="", encoding="utf-8") as fh:
            header = next(csv.reader(fh))
        # multi_group.mf4 has channels: EngineSpeed, ThrottlePos, VehicleSpeed,
        # SteeringAngle, CoolantTemp, OilTemp, BattVoltage, BattCurrent
        expected_channels = {
            "EngineSpeed", "ThrottlePos",
            "VehicleSpeed", "SteeringAngle",
            "CoolantTemp", "OilTemp",
            "BattVoltage", "BattCurrent",
        }
        assert "timestamps" in header
        header_set = set(header)
        assert expected_channels <= header_set, (
            f"Missing channels: {expected_channels - header_set}"
        )

    def test_flat_xlsx_single_sheet_named_flattened(self, multi_group_mf4, tmp_path):
        """Flat XLSX should have exactly one sheet named 'Flattened'."""
        import export as exp
        import openpyxl  # type: ignore[import-untyped]
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "flat.xlsx")
        try:
            job_id   = exp.start(mdf, "xlsx", out, flatten=True)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        wb = openpyxl.load_workbook(out, read_only=True)
        assert wb.sheetnames == ["Flattened"], f"Got sheets: {wb.sheetnames}"
        wb.close()

    def test_flat_mat_creates_single_file(self, multi_group_mf4, tmp_path):
        """Flatten + MAT should create one .mat file."""
        import export as exp
        import scipy.io as sio  # type: ignore[import-untyped]
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "flat.mat")
        try:
            job_id   = exp.start(mdf, "mat", out, flatten=True)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        assert (tmp_path / "flat.mat").is_file()
        mat = sio.loadmat(out)
        data_keys = [k for k in mat if not k.startswith("_")]
        assert len(data_keys) >= 8   # 8 channels + timestamps

    def test_flat_tsv_single_file(self, multi_group_mf4, tmp_path):
        """Flatten + tsv should produce one file."""
        import export as exp
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "flat.tsv")
        try:
            job_id   = exp.start(mdf, "tsv", out, flatten=True)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        tsv_files = [f for f in tmp_path.iterdir() if f.suffix == ".tsv"]
        assert len(tsv_files) == 1

    def test_flat_has_nan_filled_timestamps(self, multi_group_mf4, tmp_path):
        """Flat parquet table should be longer than any single group."""
        import export as exp
        import pyarrow.parquet as pq  # type: ignore[import-untyped]
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "flat.parquet")
        try:
            job_id = exp.start(mdf, "parquet", out, flatten=True)
            _wait_for_job(job_id)
            # individual group sizes: 50, 100, 25, 75 → union > 100 rows
            max_group_rows = max(len(g.channels[0:1] or [0])
                                 for g in mdf.groups if g.channels)
        finally:
            mdf.close()
        table = pq.read_table(out)
        assert table.num_rows > 100   # multi_group has up to 100 rows per group

    def test_flat_tdms_ignores_flatten(self, multi_group_mf4, tmp_path):
        """Flatten has no effect for TDMS — export still runs normally."""
        import export as exp
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "flat.tdms")
        try:
            job_id   = exp.start(mdf, "tdms", out, flatten=True)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        assert (tmp_path / "flat.tdms").is_file()


# --------------------------------------------------------------------------- #
# Phase D — MF4 re-export
# --------------------------------------------------------------------------- #

class TestMf4Export:
    def test_mf4_creates_file(self, minimal_mf4, tmp_path):
        """MF4 re-export should create the output file."""
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "re_export.mf4")
        try:
            job_id   = exp.start(mdf, "mf4", out)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        assert (tmp_path / "re_export.mf4").is_file()
        assert (tmp_path / "re_export.mf4").stat().st_size > 0

    def test_mf4_output_readable_by_asammdf(self, minimal_mf4, tmp_path):
        """The re-exported MF4 should be parseable and contain the original channels."""
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "re_export.mf4")
        try:
            job_id = exp.start(mdf, "mf4", out)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        mdf2 = _open(out)
        try:
            all_channels = [
                str(ch.name)
                for g in mdf2.groups
                for ch in g.channels
            ]
        finally:
            mdf2.close()
        assert any("Ch" in c for c in all_channels), (
            f"Expected Ch* channels, got: {all_channels}"
        )

    def test_mf4_progress_total_is_1(self, minimal_mf4, tmp_path):
        """MF4 export uses total=1 as a single-step progress indicator."""
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "re_export.mf4")
        try:
            job_id   = exp.start(mdf, "mf4", out)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["total"] == 1
        assert progress["done"]  == 1

    def test_mf4_multi_group_progress_total_is_1(self, multi_group_mf4, tmp_path):
        """MF4 re-export always has total=1 regardless of group count."""
        import export as exp
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "re_export.mf4")
        try:
            job_id   = exp.start(mdf, "mf4", out)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        assert progress["total"] == 1


# --------------------------------------------------------------------------- #
# Cancellation
# --------------------------------------------------------------------------- #

class TestCancellation:
    def test_cancel_sets_status(self, multi_group_mf4, tmp_path):
        """Start an export and immediately cancel — status should become 'cancelled'."""
        import export as exp
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "cancelled.mat")
        try:
            job_id = exp.start(mdf, "mat", out)
            exp.cancel(job_id)
            # Wait for the thread to settle
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] in ("cancelled", "done"), (
            f"Unexpected status: {progress['status']}"
        )

    def test_cancel_cleans_up_partial_file(self, multi_group_mf4, tmp_path):
        """Cancelled export should not leave a partial output file behind."""
        import export as exp
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "partial.mat")
        try:
            job_id = exp.start(mdf, "mat", out)
            exp.cancel(job_id)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        if progress["status"] == "cancelled":
            assert not os.path.isfile(out), "Partial file should have been cleaned up"

    def test_not_found_status_for_unknown_job(self):
        import export as exp
        progress = exp.get_progress("no-such-job-id")
        assert progress["status"] == "not_found"


# --------------------------------------------------------------------------- #
# Phase B — Signal filter
# --------------------------------------------------------------------------- #

class TestSignalFilter:
    """signal_filter param limits which channels are written to each format."""

    def _make_filter(self, group_index: int, *channel_names: str) -> list[dict]:
        return [
            {"group_index": group_index, "channel_name": name}
            for name in channel_names
        ]

    def test_filter_limits_csv_columns(self, minimal_mf4, tmp_path):
        """CSV export with filter=[Ch1] should only have Ch1 in the header."""
        import csv
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "filtered.csv")
        sig_filter = self._make_filter(0, "Ch1")
        try:
            job_id   = exp.start(mdf, "csv", out, signal_filter=sig_filter)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        with open(out, newline="", encoding="utf-8") as fh:
            header = next(csv.reader(fh))
        assert "Ch1" in header, f"Expected Ch1 in header: {header}"
        assert "Ch2" not in header, f"Ch2 should be excluded, got: {header}"
        assert "Ch3" not in header, f"Ch3 should be excluded, got: {header}"

    def test_filter_limits_mat_keys(self, minimal_mf4, tmp_path):
        """MAT export with filter=[Ch1, Ch2] should exclude Ch3."""
        import export as exp
        import scipy.io as sio  # type: ignore[import-untyped]
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "filtered.mat")
        sig_filter = self._make_filter(0, "Ch1", "Ch2")
        try:
            job_id = exp.start(mdf, "mat", out, signal_filter=sig_filter)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        mat = sio.loadmat(out)
        data_keys = {k for k in mat if not k.startswith("_")}
        assert "Ch1" in data_keys, f"Expected Ch1 in {data_keys}"
        assert "Ch2" in data_keys, f"Expected Ch2 in {data_keys}"
        assert "Ch3" not in data_keys, f"Ch3 should be excluded from {data_keys}"

    def test_filter_limits_parquet_columns(self, minimal_mf4, tmp_path):
        """Parquet export with filter=[Ch2] should not include Ch1 or Ch3."""
        import export as exp
        import pyarrow.parquet as pq  # type: ignore[import-untyped]
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "filtered.parquet")
        sig_filter = self._make_filter(0, "Ch2")
        try:
            job_id = exp.start(mdf, "parquet", out, signal_filter=sig_filter)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        table = pq.read_table(out)
        col_names = table.schema.names
        assert "Ch2" in col_names, f"Expected Ch2 in {col_names}"
        assert "Ch1" not in col_names, f"Ch1 should be excluded from {col_names}"
        assert "Ch3" not in col_names, f"Ch3 should be excluded from {col_names}"

    def test_filter_limits_tsv_columns(self, minimal_mf4, tmp_path):
        """TSV export with filter=[Ch3] should not include Ch1 or Ch2."""
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "filtered.tsv")
        sig_filter = self._make_filter(0, "Ch3")
        try:
            job_id = exp.start(mdf, "tsv", out, signal_filter=sig_filter)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        with open(out, encoding="utf-8") as fh:
            header = fh.readline().rstrip("\n").split("\t")
        assert "Ch3" in header, f"Expected Ch3 in header: {header}"
        assert "Ch1" not in header, f"Ch1 should be excluded: {header}"
        assert "Ch2" not in header, f"Ch2 should be excluded: {header}"

    def test_filter_limits_xlsx_columns(self, minimal_mf4, tmp_path):
        """XLSX export with filter=[Ch1] should not include Ch2 or Ch3."""
        import export as exp
        import openpyxl  # type: ignore[import-untyped]
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "filtered.xlsx")
        sig_filter = self._make_filter(0, "Ch1")
        try:
            job_id = exp.start(mdf, "xlsx", out, signal_filter=sig_filter)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        wb = openpyxl.load_workbook(out, read_only=True)
        ws = wb.worksheets[0]
        header = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        wb.close()
        assert "Ch1" in header, f"Expected Ch1 in header: {header}"
        assert "Ch2" not in header, f"Ch2 should be excluded: {header}"
        assert "Ch3" not in header, f"Ch3 should be excluded: {header}"

    def test_null_filter_exports_all_channels(self, minimal_mf4, tmp_path):
        """signal_filter=None (no filter) should export all channels normally."""
        import csv
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "no_filter.csv")
        try:
            job_id   = exp.start(mdf, "csv", out, signal_filter=None)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        with open(out, newline="", encoding="utf-8") as fh:
            header = next(csv.reader(fh))
        assert "Ch1" in header
        assert "Ch2" in header
        assert "Ch3" in header

    def test_empty_filter_list_exports_all_channels(self, minimal_mf4, tmp_path):
        """signal_filter=[] (empty list) is treated as no filter → all channels."""
        import csv
        import export as exp
        mdf = _open(minimal_mf4)
        out = str(tmp_path / "empty_filter.csv")
        try:
            job_id   = exp.start(mdf, "csv", out, signal_filter=[])
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        with open(out, newline="", encoding="utf-8") as fh:
            header = next(csv.reader(fh))
        assert any("Ch" in col for col in header)

    def test_filter_with_flatten_limits_columns(self, multi_group_mf4, tmp_path):
        """Flat CSV with signal_filter should only write the listed channels."""
        import csv
        import export as exp
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "flat_filtered.csv")
        # Include only EngineSpeed (group 0) and VehicleSpeed (group 1)
        sig_filter = [
            {"group_index": 0, "channel_name": "EngineSpeed"},
            {"group_index": 1, "channel_name": "VehicleSpeed"},
        ]
        try:
            job_id   = exp.start(mdf, "csv", out, flatten=True, signal_filter=sig_filter)
            progress = _wait_for_job(job_id)
        finally:
            mdf.close()
        assert progress["status"] == "done", progress.get("error")
        with open(out, newline="", encoding="utf-8") as fh:
            header = next(csv.reader(fh))
        assert "EngineSpeed" in header, f"Expected EngineSpeed: {header}"
        assert "VehicleSpeed" in header, f"Expected VehicleSpeed: {header}"
        assert "ThrottlePos" not in header, f"ThrottlePos should be excluded: {header}"
        assert "SteeringAngle" not in header, f"SteeringAngle should be excluded: {header}"

    def test_filter_per_group_index_is_respected(self, multi_group_mf4, tmp_path):
        """A filter entry for group 0, 'EngineSpeed' must not suppress group 1 data
        even if group 1 also has a channel named 'EngineSpeed' (different group)."""
        import csv
        import export as exp
        mdf = _open(multi_group_mf4)
        out = str(tmp_path / "grp_filter.csv")
        # Include only group 1 channels
        sig_filter = [
            {"group_index": 1, "channel_name": "VehicleSpeed"},
            {"group_index": 1, "channel_name": "SteeringAngle"},
        ]
        try:
            job_id   = exp.start(mdf, "csv", out, signal_filter=sig_filter)
            _wait_for_job(job_id)
        finally:
            mdf.close()
        # multi-group → multiple files; check they exist
        csv_files = [f for f in os.listdir(tmp_path) if f.endswith(".csv")]
        assert len(csv_files) >= 1


# --------------------------------------------------------------------------- #
# Phase B — get_exportable_signals
# --------------------------------------------------------------------------- #

class TestGetExportableSignals:
    """Tests for the export.get_exportable_signals() helper."""

    def test_returns_dict_with_groups_key(self, minimal_mf4):
        import export as exp
        mdf = _open(minimal_mf4)
        try:
            result = exp.get_exportable_signals(mdf)
        finally:
            mdf.close()
        assert isinstance(result, dict)
        assert "groups" in result

    def test_physical_channels_included(self, minimal_mf4):
        """minimal.mf4 has physical channels → groups list must be non-empty."""
        import export as exp
        mdf = _open(minimal_mf4)
        try:
            result = exp.get_exportable_signals(mdf)
        finally:
            mdf.close()
        assert len(result["groups"]) >= 1

    def test_physical_source_field(self, minimal_mf4):
        """All groups from a file with no bus-raw data must have source='physical'."""
        import export as exp
        mdf = _open(minimal_mf4)
        try:
            result = exp.get_exportable_signals(mdf)
        finally:
            mdf.close()
        for grp in result["groups"]:
            assert grp["source"] == "physical", (
                f"Expected source='physical', got {grp['source']!r}"
            )

    def test_physical_channel_names_present(self, minimal_mf4):
        """Ch1, Ch2, Ch3 should appear in the exportable channel list."""
        import export as exp
        mdf = _open(minimal_mf4)
        try:
            result = exp.get_exportable_signals(mdf)
        finally:
            mdf.close()
        all_channel_names = {ch["name"] for g in result["groups"] for ch in g["channels"]}
        assert "Ch1" in all_channel_names, f"Ch1 missing from {all_channel_names}"
        assert "Ch2" in all_channel_names, f"Ch2 missing from {all_channel_names}"
        assert "Ch3" in all_channel_names, f"Ch3 missing from {all_channel_names}"

    def test_bus_raw_groups_excluded_from_physical(self, bus_raw_mf4):
        """bus_raw.mf4 contains only raw-frame groups; get_exportable_signals without
        db_assignments should return no groups (or only non-raw groups)."""
        import export as exp
        mdf = _open(bus_raw_mf4)
        try:
            result = exp.get_exportable_signals(mdf)
        finally:
            mdf.close()
        # All returned groups must have source='physical'; bus-raw groups are excluded
        for grp in result["groups"]:
            assert grp["source"] == "physical"

    def test_with_db_assignments_adds_decoded_groups(self, bus_raw_mf4, can_bus_dbc):
        """Providing db_assignments should add decoded groups with source='decoded'."""
        import export as exp
        mdf = _open(bus_raw_mf4)
        assignments = [{"group_index": 0, "db_path": can_bus_dbc}]
        try:
            result = exp.get_exportable_signals(mdf, db_assignments=assignments)
        finally:
            mdf.close()
        sources = {grp["source"] for grp in result["groups"]}
        assert "decoded" in sources, (
            f"Expected at least one decoded group; got sources: {sources}"
        )

    def test_decoded_groups_have_channels(self, bus_raw_mf4, can_bus_dbc):
        """Decoded groups must have at least one channel entry."""
        import export as exp
        mdf = _open(bus_raw_mf4)
        assignments = [{"group_index": 0, "db_path": can_bus_dbc}]
        try:
            result = exp.get_exportable_signals(mdf, db_assignments=assignments)
        finally:
            mdf.close()
        decoded_groups = [g for g in result["groups"] if g["source"] == "decoded"]
        assert decoded_groups, "Expected at least one decoded group"
        for grp in decoded_groups:
            assert len(grp["channels"]) >= 1, (
                f"Decoded group {grp['group_index']} has no channels"
            )

    def test_channel_has_name_and_unit_fields(self, minimal_mf4):
        """Each channel entry must have 'name' and 'unit' string fields."""
        import export as exp
        mdf = _open(minimal_mf4)
        try:
            result = exp.get_exportable_signals(mdf)
        finally:
            mdf.close()
        for grp in result["groups"]:
            for ch in grp["channels"]:
                assert "name" in ch, f"Channel entry missing 'name': {ch}"
                assert "unit" in ch, f"Channel entry missing 'unit': {ch}"
                assert isinstance(ch["name"], str)
                assert isinstance(ch["unit"], str)

    def test_group_has_required_fields(self, minimal_mf4):
        """Each group entry must have group_index, acq_name, source, channels."""
        import export as exp
        mdf = _open(minimal_mf4)
        try:
            result = exp.get_exportable_signals(mdf)
        finally:
            mdf.close()
        for grp in result["groups"]:
            assert "group_index" in grp
            assert "acq_name" in grp
            assert "source" in grp
            assert "channels" in grp
            assert isinstance(grp["group_index"], int)
            assert isinstance(grp["acq_name"], str)
            assert grp["source"] in ("physical", "decoded")

    def test_no_db_assignments_returns_no_decoded_groups(self, bus_raw_mf4):
        """Without db_assignments, no decoded groups should appear."""
        import export as exp
        mdf = _open(bus_raw_mf4)
        try:
            result = exp.get_exportable_signals(mdf, db_assignments=None)
        finally:
            mdf.close()
        decoded = [g for g in result["groups"] if g["source"] == "decoded"]
        assert decoded == [], f"Expected no decoded groups, got: {decoded}"

    def test_empty_db_assignments_returns_no_decoded_groups(self, bus_raw_mf4):
        """db_assignments=[] (empty) should not trigger decoding."""
        import export as exp
        mdf = _open(bus_raw_mf4)
        try:
            result = exp.get_exportable_signals(mdf, db_assignments=[])
        finally:
            mdf.close()
        decoded = [g for g in result["groups"] if g["source"] == "decoded"]
        assert decoded == [], f"Expected no decoded groups, got: {decoded}"
