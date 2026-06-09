"""
Export jobs: MDF → .mat / .tdms / .parquet / .csv / .tsv / .xlsx.

Jobs run in background daemon threads.  Callers poll progress via
get_progress() and may request cancellation via cancel().
"""
from __future__ import annotations

import os
import re
import threading
import uuid
from typing import Any


# --------------------------------------------------------------------------- #
# Job state
# --------------------------------------------------------------------------- #

class _Job:
    __slots__ = ("total", "done", "status", "error", "_cancel", "_cleanup")

    def __init__(self, total: int) -> None:
        self.total    = total
        self.done     = 0
        self.status   = "running"         # running | done | error | cancelled
        self.error: str | None = None
        self._cancel  = threading.Event()
        self._cleanup: list[str] = []     # files to remove on cancel / error

    @property
    def cancel_requested(self) -> bool:
        return self._cancel.is_set()

    def request_cancel(self) -> None:
        self._cancel.set()


_JOBS: dict[str, _Job] = {}


# --------------------------------------------------------------------------- #
# Public API
# --------------------------------------------------------------------------- #

def _run_chunk(
    mdf: Any,
    fmt: str,
    output_path: str,
    job: Any,           # _Job or _ChunkProxy — duck-typed
    do_flatten: bool,
    mat_link_groups: bool,
    filter_set: "set[tuple[int, str]] | None",
    original_mdf: Any = None,
    chunk_start_time: Any = None,
) -> None:
    """Execute one export chunk (the whole file, or one time window when splitting).

    Called once for no-split exports, or once per time window when splitting.
    The *job* argument may be either a real ``_Job`` (no-split path) or a
    ``_ChunkProxy`` (split path) — both support the duck-type interface used
    by the individual ``_do_*`` functions.
    """
    if fmt == "mf4":
        _do_mf4(mdf, output_path, job, original_mdf=original_mdf,
                filter_set=filter_set, chunk_start_time=chunk_start_time)
    elif do_flatten:
        # Phase C — collect all groups into a flat table, then write once
        ts, cols = _build_flat_table(mdf, job, filter_set=filter_set)
        if not job.cancel_requested and cols:
            if fmt == "mat":
                _write_flat_mat(output_path, ts, cols)
            elif fmt == "parquet":
                job._cleanup.append(output_path)
                _write_flat_parquet(output_path, ts, cols)
            elif fmt in ("csv", "tsv"):
                job._cleanup.append(output_path)
                _write_flat_csv(output_path, ts, cols,
                                delimiter="," if fmt == "csv" else "\t")
            elif fmt == "xlsx":
                job._cleanup.append(output_path)
                _write_flat_xlsx(output_path, ts, cols)
    elif fmt == "mat":
        _do_mat(mdf, output_path, job,
                mat_link_groups=mat_link_groups, filter_set=filter_set)
    elif fmt == "tdms":
        _do_tdms(mdf, output_path, job, filter_set=filter_set)
    elif fmt == "parquet":
        _do_parquet(mdf, output_path, job, filter_set=filter_set)
    elif fmt == "csv":
        _do_csv(mdf, output_path, job, delimiter=",", filter_set=filter_set)
    elif fmt == "tsv":
        _do_csv(mdf, output_path, job, delimiter="\t", filter_set=filter_set)
    elif fmt == "xlsx":
        _do_xlsx(mdf, output_path, job, filter_set=filter_set)
    else:
        raise ValueError(f"unsupported format: {fmt!r}")


def start(
    mdf: Any,
    fmt: str,
    output_path: str,
    db_assignments: "list[dict] | None" = None,
    flatten: bool = False,
    mat_link_groups: bool = False,
    signal_filter: "list[dict] | None" = None,
    split_mode: str = "none",
    split_size_mb: float = 100.0,
    split_period_s: float = 60.0,
    split_first_offset_s: float = 0.0,
    source_path: str = "",
) -> str:
    """Start an export job in a background thread; return its job_id.

    If *db_assignments* is provided (a list of ``{"group_index": int, "db_path": str}``
    dicts in priority order), ``extract_bus_logging`` is applied before writing.

    If *flatten* is ``True`` and the format supports it (MAT / Parquet / CSV / TSV /
    XLSX), all channel groups are merged into a single timestamp-union table with
    NaN-filling for channels absent at a given timestamp.  TDMS and MF4 do not
    support flatten and will be exported normally.

    If *signal_filter* is provided (a list of ``{"group_index": int, "channel_name": str}``
    dicts), only the listed channels are written; all others are silently skipped.
    Omit or pass ``None`` to export all channels.

    *split_mode* may be ``"none"`` (default), ``"time"`` (fixed time windows of
    *split_period_s* seconds), or ``"size"`` (target output file size of
    *split_size_mb* MB, implemented by estimating an equivalent time period from
    the source file density).  When splitting is active the output path serves as
    a base name; each chunk is written to a file with ``_T#####s`` inserted
    before the extension, where ##### is the chunk's start time offset in seconds
    from the first sample.
    """
    # flatten only applies to tabular formats; MF4 always uses a single save() step
    do_flatten    = flatten and fmt not in ("tdms", "mf4")
    initial_total = 1 if fmt == "mf4" else len(mdf.groups)

    job_id = str(uuid.uuid4())
    job    = _Job(total=initial_total)
    _JOBS[job_id] = job

    def _run() -> None:
        try:
            # Build a fast lookup set for the signal filter (group_index, channel_name)
            filter_set: "set[tuple[int, str]] | None" = None
            if signal_filter:
                filter_set = {
                    (int(f["group_index"]), str(f["channel_name"]))
                    for f in signal_filter
                }

            active_mdf   = mdf
            original_mdf: Any = None

            if db_assignments:
                original_mdf = mdf
                active_mdf   = _build_decoded_mdf(mdf, db_assignments)
                if fmt != "mf4":
                    job.total = len(active_mdf.groups)

            # ── Split dispatch ───────────────────────────────────────────── #
            effective_split = split_mode if split_mode in ("time", "size") else "none"

            if effective_split != "none":
                t_min, t_max = _get_time_range(active_mdf)
                duration = t_max - t_min

                if duration > 0:
                    if effective_split == "size":
                        bps    = _estimate_bps(active_mdf, source_path, duration)
                        period = (split_size_mb * 1024 * 1024 / bps
                                  if bps > 0 else duration)
                        period = max(1.0, period)
                    else:
                        period = max(1.0, split_period_s)

                    windows = _split_time_windows(
                        t_min, t_max, period, split_first_offset_s
                    )

                    if len(windows) > 1:
                        job.total       = len(windows)
                        proxy           = _ChunkProxy(job)
                        # Always derive recording_start from the original file-backed
                        # MDF; extract_bus_logging may not copy the header start_time.
                        recording_start = (
                            getattr(getattr(mdf, "header", None), "start_time", None)
                            or getattr(getattr(active_mdf, "header", None), "start_time", None)
                        )

                        # When decoding is active, pre-build an (acq_name, channel_name)
                        # lookup so that the per-chunk filter_set can be rebuilt against
                        # each chunk's own decoded group indices (cut-then-decode produces
                        # independent MDFs whose group numbering may differ from the full
                        # decoded MDF).
                        name_pair_filter: "set[tuple[str, str]] | None" = None
                        if db_assignments and signal_filter:
                            name_pair_filter = {
                                (str(f.get("acq_name", "")), str(f["channel_name"]))
                                for f in signal_filter
                            }

                        for chunk_i, (w_start, w_stop) in enumerate(windows):
                            if job.cancel_requested:
                                break

                            if db_assignments:
                                # Cut-then-decode: cut the original raw MDF first so
                                # that asammdf's well-tested file-backed cut path is
                                # used, then run extract_bus_logging only on the chunk.
                                # Cutting an in-memory decoded MDF (decode-then-cut)
                                # can silently ignore the start boundary in some
                                # asammdf versions, causing every subsequent chunk to
                                # include data from t=0.
                                # time_from_zero=True rebases timestamps to 0 so each
                                # split file is self-contained.
                                raw_cut      = mdf.cut(start=w_start, stop=w_stop, whence=0, time_from_zero=True)
                                cut          = _build_decoded_mdf(raw_cut, db_assignments)
                                chunk_filter = _build_chunk_filter(cut, name_pair_filter)
                            else:
                                cut          = active_mdf.cut(start=w_start, stop=w_stop, whence=0, time_from_zero=True)
                                chunk_filter = filter_set

                            # Compute the chunk's absolute start datetime once.
                            # This is forwarded to _do_mf4 and applied to
                            # mdf_to_save just before save() — necessary because
                            # mdf.filter() creates a new MDF object that may not
                            # inherit the start_time set here on cut.
                            chunk_start_time = None
                            if recording_start is not None:
                                try:
                                    from datetime import timedelta as _td
                                    chunk_start_time = recording_start + _td(seconds=w_start)
                                    cut.header.start_time = chunk_start_time
                                except Exception:
                                    pass

                            suffix     = _absolute_suffix(w_start, recording_start)
                            chunk_path = _insert_suffix(output_path, suffix)
                            _run_chunk(cut, fmt, chunk_path, proxy,
                                       do_flatten, mat_link_groups,
                                       chunk_filter, original_mdf,
                                       chunk_start_time=chunk_start_time)
                            job.done = chunk_i + 1

                        to_delete = job._cleanup if job._cleanup else [output_path]
                        if job.cancel_requested:
                            for p in to_delete:
                                _delete(p)
                        else:
                            job.status = "done"
                        return
            # Fall through: no split (or duration ≤ 0 / single window only)

            # ── Single-file path ─────────────────────────────────────────── #
            _run_chunk(active_mdf, fmt, output_path, job,
                       do_flatten, mat_link_groups, filter_set, original_mdf)

            to_delete = job._cleanup if job._cleanup else [output_path]
            if job.cancel_requested:
                for p in to_delete:
                    _delete(p)
            else:
                job.status = "done"

        except Exception as exc:   # noqa: BLE001
            job.error  = str(exc)
            job.status = "error"
            for p in (job._cleanup if job._cleanup else [output_path]):
                _delete(p)

    threading.Thread(target=_run, daemon=True).start()
    return job_id


def preview_bus_decoding(mdf: Any, db_assignments: list[dict]) -> list[dict]:
    """Lightweight DB-preview scan (no full decode).

    For each ``{"group_index": int, "db_path": str}`` entry, count how many
    messages in the DB have IDs that appear in the group's raw CAN frames.
    When the group has no ``CAN_DataFrame.ID`` channel the full DB is reported.

    Returns a list of result dicts in the same order as the input.
    """
    db_cache:  dict[str, Any]       = {}   # db_path → canmatrix CanMatrix
    id_cache:  dict[int, "set[int] | None"] = {}   # group_index → CAN ID set
    results:   list[dict]           = []

    for assignment in db_assignments:
        group_index = int(assignment["group_index"])
        db_path     = str(assignment["db_path"])
        res: dict = {
            "group_index":      group_index,
            "db_path":          db_path,
            "matched_messages": 0,
            "signal_count":     0,
            "error":            None,
        }

        try:
            if db_path not in db_cache:
                db_cache[db_path] = _load_db_matrix(db_path)
            db = db_cache[db_path]

            db_msg_map = _db_message_map(db)  # {arb_id: signal_count}

            if group_index not in id_cache:
                id_cache[group_index] = _get_group_can_ids(mdf, group_index)
            group_ids = id_cache[group_index]

            if group_ids is None:
                # Can't read IDs from this group — report all DB entries as matched.
                matched = db_msg_map
            else:
                matched = {aid: sc for aid, sc in db_msg_map.items()
                           if aid in group_ids}

            res["matched_messages"] = len(matched)
            res["signal_count"]     = sum(matched.values())

        except Exception as exc:  # noqa: BLE001
            res["error"] = str(exc)

        results.append(res)

    return results


def get_progress(job_id: str) -> dict[str, Any]:
    job = _JOBS.get(job_id)
    if job is None:
        return {"status": "not_found", "done": 0, "total": 0, "error": None}
    return {
        "status": job.status,
        "done":   job.done,
        "total":  job.total,
        "error":  job.error,
    }


def cancel(job_id: str) -> None:
    job = _JOBS.get(job_id)
    if job and job.status == "running":
        job.request_cancel()
        job.status = "cancelled"


# --------------------------------------------------------------------------- #
# .mat export
# --------------------------------------------------------------------------- #

def _do_mat(
    mdf: Any,
    output_path: str,
    job: _Job,
    mat_link_groups: bool = False,
    filter_set: "set[tuple[int, str]] | None" = None,
) -> None:
    """Write a ``.mat`` v5 file.

    Each non-empty channel group produces a timestamp vector named ``t1``,
    ``t2``, … in the output file.  Channel variables are named with the
    standard ``_mat_var`` sanitiser.

    When *mat_link_groups* is ``True`` the data-vector name for every channel
    in group *i* is suffixed with the matching time-vector label (e.g.
    ``EngineSpeed_t1``), making the channel-to-time-axis association explicit
    when the file is loaded in MATLAB.
    """
    try:
        import scipy.io as sio  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError(f"scipy is required for .mat export: {exc}") from exc

    import numpy as np

    mat_data: dict[str, Any] = {}
    seen:     dict[str, int] = {}   # uniqueness counter for _mat_var

    # ── Phase 1: pre-register time-vector labels ─────────────────────────────
    # Assign t1, t2, … to every group that has channel entries and reserve
    # these names in `seen` so no channel variable can accidentally steal them.
    group_time_label: dict[int, str] = {}  # group_index → actual MATLAB var name
    t_counter = 1
    for i, group in enumerate(mdf.groups):
        if group.channels:
            group_time_label[i] = _mat_var(f"t{t_counter}", seen)
            t_counter += 1

    # ── Phase 2: iterate groups, collect data and timestamps ─────────────────
    for i, group in enumerate(mdf.groups):
        if job.cancel_requested:
            return

        time_label    = group_time_label.get(i)   # e.g. "t1", "t2", …
        timestamps_arr: "Any | None" = None
        found_numeric = False

        for ch in group.channels:
            name = str(ch.name or "")
            if not name:
                continue
            if filter_set is not None and (i, name) not in filter_set:
                continue
            try:
                sig     = mdf.get(name, group=i, raw=False,
                                  ignore_invalidation_bits=True)
                samples = sig.samples
                if not (hasattr(samples, "dtype")
                        and np.issubdtype(samples.dtype, np.number)):
                    continue
                if timestamps_arr is None and sig.timestamps is not None:
                    timestamps_arr = sig.timestamps

                if mat_link_groups and time_label:
                    var_name = _mat_var(f"{name}_{time_label}", seen)
                else:
                    var_name = _mat_var(name, seen)
                mat_data[var_name] = samples
                found_numeric = True
            except Exception:  # noqa: BLE001
                pass

        # Export the timestamp vector only when there is numeric data for it
        if found_numeric and timestamps_arr is not None and time_label:
            mat_data[time_label] = timestamps_arr

        job.done = i + 1

    sio.savemat(output_path, mat_data, do_compression=True)


# --------------------------------------------------------------------------- #
# .tdms export
# --------------------------------------------------------------------------- #

def _do_tdms(mdf: Any, output_path: str, job: _Job,
             filter_set: "set[tuple[int, str]] | None" = None) -> None:
    try:
        from nptdms import TdmsWriter, ChannelObject  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError(f"nptdms is required for .tdms export: {exc}") from exc

    import numpy as np

    with TdmsWriter(output_path) as writer:
        for i, group in enumerate(mdf.groups):
            if job.cancel_requested:
                return

            cg       = group.channel_group
            grp_name = (str(getattr(cg, "acq_name", "") or "").strip()
                        or f"Group_{i}")

            channels: list[Any] = []
            for ch in group.channels:
                name = str(ch.name or "")
                if not name:
                    continue
                if filter_set is not None and (i, name) not in filter_set:
                    continue
                try:
                    sig     = mdf.get(name, group=i, raw=False,
                                      ignore_invalidation_bits=True)
                    samples = sig.samples
                    if not (hasattr(samples, "dtype")
                            and np.issubdtype(samples.dtype, np.number)
                            and samples.ndim == 1):
                        continue
                    channels.append(ChannelObject(grp_name, name, samples))
                except Exception:  # noqa: BLE001
                    pass

            if channels:
                writer.write_segment(channels)

            job.done = i + 1


# --------------------------------------------------------------------------- #
# .parquet export
# --------------------------------------------------------------------------- #

def _do_parquet(mdf: Any, output_path: str, job: _Job,
                filter_set: "set[tuple[int, str]] | None" = None) -> None:
    """
    Write one Parquet file per non-empty channel group.

    • Single non-empty group  → written to *output_path* exactly.
    • Multiple non-empty groups → written to
        ``{stem}_g{i:02d}_{safe_acq_name}.parquet``
      where *stem* is *output_path* with the ``.parquet`` suffix stripped.

    Each file contains a ``timestamps`` column (float64 seconds) followed by
    one column per channel.  Columns that cannot be converted to an Arrow
    array are silently skipped.  All arrays are truncated to the shortest
    length in the group to guarantee a rectangular table.
    """
    try:
        import pyarrow as pa            # type: ignore[import-untyped]
        import pyarrow.parquet as pq    # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError(
            f"pyarrow is required for Parquet export: {exc}"
        ) from exc

    from pathlib import Path

    base = Path(output_path)
    stem = base.with_suffix("")   # strip .parquet for multi-file naming

    non_empty_count = sum(1 for g in mdf.groups if g.channels)
    single_file     = non_empty_count <= 1

    for i, group in enumerate(mdf.groups):
        if job.cancel_requested:
            return

        if not group.channels:
            job.done = i + 1
            continue

        cg       = group.channel_group
        grp_name = (str(getattr(cg, "acq_name", "") or "").strip()
                    or f"group_{i}")

        # ── collect raw numpy arrays ────────────────────────────────────────
        timestamps_arr = None
        columns: dict[str, Any] = {}

        for ch in group.channels:
            ch_name = str(ch.name or "")
            if not ch_name:
                continue
            if filter_set is not None and (i, ch_name) not in filter_set:
                continue
            try:
                sig = mdf.get(ch_name, group=i, raw=False,
                              ignore_invalidation_bits=True)
                if timestamps_arr is None and sig.timestamps is not None:
                    timestamps_arr = sig.timestamps
                columns[ch_name] = sig.samples
            except Exception:  # noqa: BLE001
                pass

        if not columns:
            job.done = i + 1
            continue

        # ── align lengths ───────────────────────────────────────────────────
        all_arrs = list(columns.values())
        if timestamps_arr is not None:
            all_arrs.append(timestamps_arr)
        min_len = min(len(a) for a in all_arrs)

        # ── build Arrow table ───────────────────────────────────────────────
        arrow_cols: dict[str, Any] = {}
        if timestamps_arr is not None:
            arrow_cols["timestamps"] = pa.array(
                timestamps_arr[:min_len], type=pa.float64()
            )
        for col_name, arr in columns.items():
            try:
                arrow_cols[col_name] = pa.array(arr[:min_len])
            except Exception:  # noqa: BLE001
                pass  # skip unconvertible columns (e.g. structured dtypes)

        if not arrow_cols:
            job.done = i + 1
            continue

        table = pa.table(arrow_cols)

        # ── resolve output path ─────────────────────────────────────────────
        if single_file:
            out_file = str(base)
        else:
            safe = re.sub(r"[^\w\-]", "_", grp_name)[:40].strip("_") or f"g{i}"
            out_file = str(Path(f"{stem}_g{i:02d}_{safe}.parquet"))

        job._cleanup.append(out_file)
        pq.write_table(table, out_file, compression="snappy")

        job.done = i + 1


# --------------------------------------------------------------------------- #
# .csv / .tsv export
# --------------------------------------------------------------------------- #

def _do_csv(mdf: Any, output_path: str, job: _Job, delimiter: str = ",",
            filter_set: "set[tuple[int, str]] | None" = None) -> None:
    """
    Write one delimited-text file per non-empty channel group.

    • Single non-empty group  → written to *output_path* exactly.
    • Multiple non-empty groups → written to
        ``{stem}_g{i:02d}_{safe_acq_name}{ext}``
      where *stem* is *output_path* with the extension stripped.

    Each file has a header row followed by one row per sample.
    Columns: ``timestamps`` (seconds, float) then one column per numeric channel.
    """
    import csv as _csv

    import numpy as np
    from pathlib import Path

    base = Path(output_path)
    ext  = base.suffix          # .csv or .tsv
    stem = base.with_suffix("")

    non_empty_count = sum(1 for g in mdf.groups if g.channels)
    single_file     = non_empty_count <= 1

    for i, group in enumerate(mdf.groups):
        if job.cancel_requested:
            return

        if not group.channels:
            job.done = i + 1
            continue

        cg       = group.channel_group
        grp_name = (str(getattr(cg, "acq_name", "") or "").strip()
                    or f"group_{i}")

        # ── collect numeric arrays ──────────────────────────────────────────
        timestamps_arr = None
        columns: dict[str, Any] = {}

        for ch in group.channels:
            ch_name = str(ch.name or "")
            if not ch_name:
                continue
            if filter_set is not None and (i, ch_name) not in filter_set:
                continue
            try:
                sig = mdf.get(ch_name, group=i, raw=False,
                              ignore_invalidation_bits=True)
                # Skip non-numeric channels and 2D arrays (e.g. DataBytes).
                if not (hasattr(sig.samples, "dtype")
                        and np.issubdtype(sig.samples.dtype, np.number)
                        and sig.samples.ndim == 1):
                    continue
                if timestamps_arr is None and sig.timestamps is not None:
                    timestamps_arr = sig.timestamps
                columns[ch_name] = sig.samples
            except Exception:  # noqa: BLE001
                pass

        if not columns:
            job.done = i + 1
            continue

        # ── align lengths ───────────────────────────────────────────────────
        all_arrs = list(columns.values())
        if timestamps_arr is not None:
            all_arrs.append(timestamps_arr)
        min_len = min(len(a) for a in all_arrs)

        # ── resolve output path ─────────────────────────────────────────────
        if single_file:
            out_file = str(base)
        else:
            safe = re.sub(r"[^\w\-]", "_", grp_name)[:40].strip("_") or f"g{i}"
            out_file = str(Path(f"{stem}_g{i:02d}_{safe}{ext}"))

        job._cleanup.append(out_file)

        # ── write ───────────────────────────────────────────────────────────
        col_names = list(columns.keys())
        col_arrs  = [columns[n][:min_len] for n in col_names]
        ts_slice  = timestamps_arr[:min_len] if timestamps_arr is not None else None

        with open(out_file, "w", newline="", encoding="utf-8") as fh:
            writer = _csv.writer(fh, delimiter=delimiter)
            header = (["timestamps"] if ts_slice is not None else []) + col_names
            writer.writerow(header)
            for j in range(min_len):
                row = (([float(ts_slice[j])] if ts_slice is not None else [])
                       + [float(a[j]) for a in col_arrs])
                writer.writerow(row)

        job.done = i + 1


# --------------------------------------------------------------------------- #
# .xlsx export
# --------------------------------------------------------------------------- #

def _do_xlsx(mdf: Any, output_path: str, job: _Job,
             filter_set: "set[tuple[int, str]] | None" = None) -> None:
    """
    Write a single Excel workbook with one worksheet per non-empty channel group.

    Sheet names are derived from the acquisition name (max 31 chars, special
    characters replaced with underscores — Excel's sheet-name rules).
    Each sheet has a header row followed by one row per sample.
    Columns: ``timestamps`` (seconds, float) then one column per numeric channel.
    """
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is required for .xlsx export: {exc}") from exc

    import numpy as np

    wb = openpyxl.Workbook(write_only=True)
    job._cleanup.append(output_path)

    for i, group in enumerate(mdf.groups):
        if job.cancel_requested:
            return

        if not group.channels:
            job.done = i + 1
            continue

        cg       = group.channel_group
        grp_name = (str(getattr(cg, "acq_name", "") or "").strip()
                    or f"Group_{i}")
        # Excel sheet name rules: ≤31 chars, no \ / * ? [ ] :
        sheet_name = re.sub(r'[\\/*?\[\]:]', "_", grp_name)[:31].strip() or f"Group_{i}"

        # ── collect numeric arrays ──────────────────────────────────────────
        timestamps_arr = None
        columns: dict[str, Any] = {}

        for ch in group.channels:
            ch_name = str(ch.name or "")
            if not ch_name:
                continue
            if filter_set is not None and (i, ch_name) not in filter_set:
                continue
            try:
                sig = mdf.get(ch_name, group=i, raw=False,
                              ignore_invalidation_bits=True)
                # Skip non-numeric channels and 2D arrays (e.g. DataBytes).
                if not (hasattr(sig.samples, "dtype")
                        and np.issubdtype(sig.samples.dtype, np.number)
                        and sig.samples.ndim == 1):
                    continue
                if timestamps_arr is None and sig.timestamps is not None:
                    timestamps_arr = sig.timestamps
                columns[ch_name] = sig.samples
            except Exception:  # noqa: BLE001
                pass

        if not columns:
            job.done = i + 1
            continue

        # ── align lengths ───────────────────────────────────────────────────
        all_arrs = list(columns.values())
        if timestamps_arr is not None:
            all_arrs.append(timestamps_arr)
        min_len = min(len(a) for a in all_arrs)

        # ── write sheet ─────────────────────────────────────────────────────
        ws = wb.create_sheet(title=sheet_name)

        col_names = list(columns.keys())
        col_arrs  = [columns[n][:min_len] for n in col_names]
        ts_slice  = timestamps_arr[:min_len] if timestamps_arr is not None else None

        header = (["timestamps"] if ts_slice is not None else []) + col_names
        ws.append(header)

        for j in range(min_len):
            row = (([float(ts_slice[j])] if ts_slice is not None else [])
                   + [float(a[j]) for a in col_arrs])
            ws.append(row)

        job.done = i + 1

    wb.save(output_path)


# --------------------------------------------------------------------------- #
# .mf4 re-export (Phase D)
# --------------------------------------------------------------------------- #

# Raw-bus channel names — presence of any of these means the group is still a
# genuine bus-frame group and must keep its BUS_EVENT flag.
_RAW_BUS_CHANNEL_NAMES: frozenset[str] = frozenset({
    "CAN_DataFrame", "CAN_RemoteFrame", "CAN_ErrorFrame", "CAN_DataFrame_Raw",
    "CAN_DataFrame.BRS", "CAN_DataFrame.EDL",
    "LIN_Frame", "LIN_SyncError", "LIN_ReceiveError", "LIN_WakeUp",
    "LIN_Checksum_Error",
    "FlexRay_Frame", "FlexRay_RxError", "FlexRay_TxError", "FlexRay_PDU",
    "Ethernet_Frame", "Ethernet_RxError", "Ethernet_TxError",
    "MOST_Frame",
})


def _clear_decoded_bus_event_flags(mdf_obj: Any) -> None:
    """Clear the MDF4 BUS_EVENT channel-group flag (cg_flags bit 0x02) from
    every group that has the flag set but no longer contains raw-bus channels.

    ``extract_bus_logging`` replaces raw-frame groups with decoded-signal
    groups at the same indices but leaves the original ``cg_flags`` intact,
    which causes downstream viewers to display the decoded groups as raw-frame
    groups.  Clearing the flag (while leaving the group untouched in all other
    respects) corrects that classification.

    Groups that still hold raw-bus channel names (e.g. undecoded frame groups
    for which no DB was assigned) keep their flag unchanged.
    """
    for grp in mdf_obj.groups:
        cg = grp.channel_group
        if not (hasattr(cg, "flags") and (int(cg.flags) & 0x02)):
            continue  # BUS_EVENT flag not set — nothing to fix
        ch_names = {ch.name for ch in grp.channels}
        if not (ch_names & _RAW_BUS_CHANNEL_NAMES):
            # No raw-bus channels remain — this is a decoded group
            cg.flags = int(cg.flags) & ~0x02


def _do_mf4(
    mdf: Any,
    output_path: str,
    job: _Job,
    original_mdf: Any = None,
    filter_set: "set[tuple[int, str]] | None" = None,
    chunk_start_time: Any = None,
) -> None:
    """Re-export *mdf* to an MF4 file using ``MDF.save()``.

    Progress: total = 1; done = 1 after ``save()`` returns.

    When *original_mdf* is provided (i.e., *mdf* came from
    ``extract_bus_logging``), the original HD metadata fields are copied
    onto the decoded MDF header before saving so provenance is preserved.

    When *filter_set* is provided, only the listed channels are kept via
    ``MDF.filter()`` before saving.
    """
    job.total = 1
    job.done  = 0

    if original_mdf is not None:
        try:
            src_hdr = original_mdf.header
            dst_hdr = mdf.header
            for attr in ("author", "department", "project", "subject", "comment"):
                val = getattr(src_hdr, attr, None)
                if val is not None:
                    try:
                        setattr(dst_hdr, attr, val)
                    except Exception:  # noqa: BLE001
                        pass
        except Exception:  # noqa: BLE001
            pass

    mdf_to_save = mdf
    if filter_set is not None:
        # Build (name, group, channel_index) 3-tuples so asammdf resolves each
        # channel via a direct lookup instead of the ambiguity-check path.
        # 2-tuple (name, group) still triggers "Multiple occurrences" +
        # logger.exception() (→ "[sidecar] NoneType: None") when the same name
        # appears more than once in a group (e.g. decoded bus signals).  The
        # 3-tuple path skips that check entirely: it just verifies
        # (group, index) in channels_db[name] and returns immediately.
        channels_db = getattr(getattr(mdf, "_mdf", None), "channels_db", {})
        channel_specs: list[tuple[str, int, int]] = []
        for grp_idx, ch_name in filter_set:
            ch_idx = next(
                (ci for gi, ci in channels_db.get(ch_name, ()) if gi == grp_idx),
                None,
            )
            if ch_idx is not None:
                channel_specs.append((ch_name, grp_idx, ch_idx))
        try:
            mdf_to_save = mdf.filter(channel_specs) if channel_specs else mdf
        except Exception:  # noqa: BLE001
            mdf_to_save = mdf  # fallback: save unfiltered

    # When bus decoding was applied, clear the BUS_EVENT flag (bit 0x02) from
    # channel groups that no longer contain raw-bus channels.  asammdf's
    # extract_bus_logging keeps the original cg_flags on the decoded groups, so
    # without this step MDF viewers mark them as raw-frame groups even though
    # they now hold physical signals.
    if original_mdf is not None:
        _clear_decoded_bus_event_flags(mdf_to_save)

    # For split exports, mdf.filter() above creates a new MDF object and may
    # not propagate the start_time we set on the cut MDF.  Apply the correct
    # chunk start_time to mdf_to_save just before saving so the saved file
    # always has the right header metadata.
    if chunk_start_time is not None:
        try:
            mdf_to_save.header.start_time = chunk_start_time
        except Exception:  # noqa: BLE001
            pass

    job._cleanup.append(output_path)
    mdf_to_save.save(output_path, overwrite=True)
    job.done = 1


# --------------------------------------------------------------------------- #
# Flatten helpers (Phase C)
# --------------------------------------------------------------------------- #

def _build_flat_table(
    mdf: Any,
    job: _Job,
    filter_set: "set[tuple[int, str]] | None" = None,
) -> "tuple[Any, dict[str, Any]]":
    """Collect all numeric channels across all groups into a flat timestamp-union table.

    Returns ``(timestamps, columns)`` where *timestamps* is a sorted float64
    array of every unique sample time across all groups and *columns* is an
    ``{name: float64_array}`` dict with ``nan`` at timestamps where a channel
    has no sample.

    Progress: ``job.done`` is incremented to ``i + 1`` after each group.
    """
    import numpy as np

    seen_names: set[str]  = set()
    group_data: list[Any] = []   # [(float64_ts, {col_name: float64_arr})]

    for i, group in enumerate(mdf.groups):
        if job.cancel_requested:
            return np.array([]), {}

        ts_arr:  "Any | None"    = None
        columns: dict[str, Any]  = {}

        for ch in group.channels:
            ch_name = str(ch.name or "")
            if not ch_name:
                continue
            if filter_set is not None and (i, ch_name) not in filter_set:
                continue
            try:
                sig = mdf.get(ch_name, group=i, raw=False,
                              ignore_invalidation_bits=True)
                # Skip non-numeric channels and 2D arrays (e.g. DataBytes).
                if not (hasattr(sig.samples, "dtype")
                        and np.issubdtype(sig.samples.dtype, np.number)
                        and sig.samples.ndim == 1):
                    continue
                if ts_arr is None and sig.timestamps is not None:
                    ts_arr = sig.timestamps.astype(np.float64)
                columns[ch_name] = sig.samples
            except Exception:  # noqa: BLE001
                pass

        if ts_arr is None or not columns:
            job.done = i + 1
            continue

        # Align all arrays in this group to the shortest length
        min_len = min(len(ts_arr), min(len(a) for a in columns.values()))

        # Deduplicate column names across groups by appending _g{i}
        aligned: dict[str, Any] = {}
        for ch_name, arr in columns.items():
            key = ch_name
            if key in seen_names:
                key = f"{ch_name}_g{i}"
                cnt = 2
                while key in seen_names:
                    key = f"{ch_name}_g{i}_{cnt}"
                    cnt += 1
            seen_names.add(key)
            aligned[key] = arr[:min_len].astype(np.float64)

        group_data.append((ts_arr[:min_len], aligned))
        job.done = i + 1

    if not group_data:
        return np.array([]), {}

    # Build sorted union of all group timestamps
    all_ts = np.unique(np.concatenate([ts for ts, _ in group_data]))

    # NaN-fill each group's channels at missing union timestamps
    merged: dict[str, Any] = {}
    for g_ts, g_cols in group_data:
        indices = np.searchsorted(all_ts, g_ts)
        for col_name, arr in g_cols.items():
            col = np.full(len(all_ts), np.nan, dtype=np.float64)
            col[indices] = arr
            merged[col_name] = col

    return all_ts, merged


def _write_flat_mat(
    output_path: str,
    timestamps: Any,
    columns: dict[str, Any],
) -> None:
    try:
        import scipy.io as sio  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError(f"scipy is required for .mat export: {exc}") from exc

    seen: dict[str, int] = {}
    mat_data: dict[str, Any] = {}
    if len(timestamps):
        mat_data[_mat_var("timestamps", seen)] = timestamps
    for col_name, arr in columns.items():
        mat_data[_mat_var(col_name, seen)] = arr
    sio.savemat(output_path, mat_data, do_compression=True)


def _write_flat_parquet(
    output_path: str,
    timestamps: Any,
    columns: dict[str, Any],
) -> None:
    try:
        import pyarrow as pa            # type: ignore[import-untyped]
        import pyarrow.parquet as pq    # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError(
            f"pyarrow is required for Parquet export: {exc}"
        ) from exc

    arrow_cols: dict[str, Any] = {}
    if len(timestamps):
        arrow_cols["timestamps"] = pa.array(timestamps, type=pa.float64())
    for col_name, arr in columns.items():
        try:
            arrow_cols[col_name] = pa.array(arr)
        except Exception:  # noqa: BLE001
            pass
    if arrow_cols:
        pq.write_table(pa.table(arrow_cols), output_path, compression="snappy")


def _write_flat_csv(
    output_path: str,
    timestamps: Any,
    columns: dict[str, Any],
    delimiter: str = ",",
) -> None:
    import csv as _csv

    col_names = list(columns.keys())
    has_ts    = len(timestamps) > 0
    n         = len(timestamps) if has_ts else (
        max((len(a) for a in columns.values()), default=0)
    )

    with open(output_path, "w", newline="", encoding="utf-8") as fh:
        writer = _csv.writer(fh, delimiter=delimiter)
        header = (["timestamps"] if has_ts else []) + col_names
        writer.writerow(header)
        col_arrs = [columns[name] for name in col_names]
        for j in range(n):
            row = (
                [float(timestamps[j])] if has_ts else []
            ) + [float(a[j]) for a in col_arrs]
            writer.writerow(row)


def _write_flat_xlsx(
    output_path: str,
    timestamps: Any,
    columns: dict[str, Any],
) -> None:
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is required for .xlsx export: {exc}") from exc

    import math

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title="Flattened")

    col_names = list(columns.keys())
    has_ts    = len(timestamps) > 0
    header    = (["timestamps"] if has_ts else []) + col_names
    ws.append(header)

    n        = len(timestamps) if has_ts else (
        max((len(a) for a in columns.values()), default=0)
    )
    col_arrs = [columns[name] for name in col_names]

    for j in range(n):
        row = (
            [float(timestamps[j])] if has_ts else []
        ) + [
            None if math.isnan(a[j]) else float(a[j])
            for a in col_arrs
        ]
        ws.append(row)

    wb.save(output_path)


# --------------------------------------------------------------------------- #
# Channel-filter helpers (Phase B)
# --------------------------------------------------------------------------- #

def get_exportable_signals(
    mdf: Any,
    db_assignments: "list[dict] | None" = None,
) -> dict:
    """Return all exportable channels grouped by channel group.

    Without *db_assignments* (or empty list): returns only physical (non-raw-frame)
    channel groups, each with ``source="physical"``.

    With *db_assignments*: additionally runs ``extract_bus_logging`` and returns
    all decoded groups with ``source="decoded"``.  Physical groups are still
    included.  Decoded group indices are their positions in the decoded MDF
    (0, 1, 2, …), which are independent of the original MDF's group indices.

    Return schema::

        {
            "groups": [
                {
                    "group_index": int,
                    "acq_name":    str,
                    "source":      "physical" | "decoded",
                    "channels":    [{"name": str, "unit": str}],
                }
            ]
        }
    """
    import metadata as _meta

    groups_out: list[dict] = []

    # ── Physical channels from non-bus groups ────────────────────────────────
    for i, group in enumerate(mdf.groups):
        if _meta.group_bus_type(group) is not None:
            continue  # skip raw-frame groups

        cg = group.channel_group
        acq_name = str(getattr(cg, "acq_name", "") or "").strip() or f"Group {i}"

        channels_out = [
            {"name": str(ch.name or ""), "unit": str(ch.unit or "").strip()}
            for ch in group.channels
            if ch.name
        ]
        if channels_out:
            groups_out.append({
                "group_index": i,
                "acq_name":    acq_name,
                "source":      "physical",
                "channels":    channels_out,
            })

    # ── Decoded channels when db_assignments are provided ───────────────────
    # extract_bus_logging produces a brand-new MDF containing *only* the
    # decoded groups — one per matched DBC message — appended in order.
    # Its group indices (0, 1, 2, …) are independent of the original MDF's
    # indices.  We must iterate the decoded MDF directly and use its actual
    # indices; filtering by the original bus-raw indices would discard most
    # decoded groups.
    if db_assignments:
        decoded_mdf = _build_decoded_mdf(mdf, db_assignments)
        for i, group in enumerate(decoded_mdf.groups):
            cg = group.channel_group
            acq_name = str(getattr(cg, "acq_name", "") or "").strip() or f"Group {i}"

            channels_out = [
                {"name": str(ch.name or ""), "unit": str(ch.unit or "").strip()}
                for ch in group.channels
                if ch.name
            ]
            if channels_out:
                groups_out.append({
                    "group_index": i,
                    "acq_name":    acq_name,
                    "source":      "decoded",
                    "channels":    channels_out,
                })

    return {"groups": groups_out}


# --------------------------------------------------------------------------- #
# Bus-decoding helpers (Phase A)
# --------------------------------------------------------------------------- #

def _load_db_matrix(db_path: str) -> Any:
    """Load a ``.dbc`` or ``.arxml`` file.  Returns a canmatrix ``CanMatrix``."""
    try:
        import canmatrix.formats  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError(
            f"canmatrix is required for bus decoding: {exc}"
        ) from exc
    db_dict = canmatrix.formats.loadp(db_path)
    if not db_dict:
        raise RuntimeError(f"No network found in database file: {db_path!r}")
    return next(iter(db_dict.values()))


def _db_message_map(db: Any) -> dict[int, int]:
    """Return ``{normalised_arb_id: signal_count}`` for all frames in *db*.

    canmatrix uses ``CanMatrix.frames`` (a list of ``Frame`` objects).
    The ``arbitration_id`` field is a ``canmatrix.ArbitrationId`` object
    (with a ``.id`` int attribute) in recent versions; older versions store
    a plain int with bit 31 set for extended frames.
    """
    result: dict[int, int] = {}
    frames = getattr(db, "frames", None) or getattr(db, "messages", [])
    for frame in frames:
        arb = frame.arbitration_id
        msg_id = int(getattr(arb, "id", arb)) & 0x1FFF_FFFF
        result[msg_id] = len(list(frame.signals))
    return result


def _get_group_can_ids(mdf: Any, group_index: int) -> "set[int] | None":
    """Read unique 29-bit CAN IDs from a raw group.  Returns *None* on failure."""
    try:
        sig = mdf.get(
            "CAN_DataFrame.ID",
            group=group_index,
            raw=True,
            ignore_invalidation_bits=True,
        )
        return {int(x) & 0x1FFF_FFFF for x in sig.samples.tolist()}
    except Exception:  # noqa: BLE001
        return None


# asammdf extract_bus_logging key per our metadata bus-type label
_BUS_TYPE_TO_EBL_KEY: dict[str, str] = {
    "CAN":      "CAN",
    "CAN FD":   "CAN",   # CAN FD frames are decoded with CAN databases
    "LIN":      "LIN",
    "FlexRay":  "FLEXRAY",
    "MOST":     "MOST",
    "Ethernet": "Ethernet",
    "K-Line":   "K_LINE",
}


def _get_group_bus_channel(mdf: Any, group_index: int) -> int:
    """Return the bus channel number stored in a raw-bus group.

    Reads the first sample of ``CAN_DataFrame.BusChannel`` (or its LIN / FlexRay
    equivalent via a generic channel name search) so that the correct per-channel
    entry is passed to ``asammdf.extract_bus_logging``.

    Returns 0 when the channel number cannot be determined.  In asammdf's
    ``extract_bus_logging`` convention, channel 0 is the "all-channels" wildcard,
    so falling back to 0 preserves the previous behaviour for files that do not
    carry an explicit BusChannel signal.
    """
    _BUS_CHANNEL_NAMES = (
        "CAN_DataFrame.BusChannel",
        "LIN_Frame.BusChannel",
        "FlexRay_Frame.BusChannel",
        "Ethernet_Frame.BusChannel",
    )
    for ch_name in _BUS_CHANNEL_NAMES:
        try:
            sig = mdf.get(
                ch_name, group=group_index, raw=True,
                ignore_invalidation_bits=True,
            )
            if sig.samples is not None and len(sig.samples) > 0:
                return int(sig.samples[0])
        except Exception:  # noqa: BLE001
            continue
    return 0


def _build_decoded_mdf(mdf: Any, db_assignments: list[dict]) -> Any:
    """Apply ``MDF.extract_bus_logging()`` using the given ordered DB assignments.

    Builds the ``database_files`` dict expected by asammdf — a mapping from bus
    type key (e.g. ``"CAN"``) to an ordered list of ``(db_path, bus_channel)``
    tuples.  The bus_channel for each entry is read from the group's
    ``CAN_DataFrame.BusChannel`` signal so that databases are scoped to the
    correct physical channel.  Using the actual channel number (rather than the
    wildcard 0) prevents a DBC assigned to channel 1 from also decoding frames
    that belong to channel 2 (and vice-versa) when both channels carry the same
    message IDs with different signal definitions.

    Returns the decoded MDF object (a new instance; the original is unmodified).
    """
    import metadata as _meta  # sidecar root is on sys.path at runtime

    # ebl_key → ordered list of (db_path, bus_channel) tuples, deduplicated
    ebl_dbs: dict[str, list[tuple[str, int]]]  = {}
    seen:    dict[str, set[tuple[str, int]]]   = {}

    for assignment in db_assignments:
        group_index = int(assignment["group_index"])
        db_path     = str(assignment["db_path"])

        if group_index >= len(mdf.groups):
            continue

        group    = mdf.groups[group_index]
        bus_type = _meta.group_bus_type(group) or ""
        ebl_key  = _BUS_TYPE_TO_EBL_KEY.get(bus_type)
        if not ebl_key:
            continue

        # Resolve the bus channel for this group so asammdf can scope the
        # database to the right physical channel.  Falls back to 0 (wildcard)
        # when the signal is absent.
        bus_channel = _get_group_bus_channel(mdf, group_index)
        entry       = (db_path, bus_channel)

        if ebl_key not in ebl_dbs:
            ebl_dbs[ebl_key] = []
            seen[ebl_key]    = set()

        if entry not in seen[ebl_key]:
            ebl_dbs[ebl_key].append(entry)
            seen[ebl_key].add(entry)

    if not ebl_dbs:
        return mdf

    # asammdf expects {"CAN": [(path, bus_channel), ...], ...}
    database_files: dict[str, list[tuple[str, int]]] = dict(ebl_dbs)

    return mdf.extract_bus_logging(database_files=database_files)


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def _mat_var(name: str, seen: dict[str, int]) -> str:
    """Unique MATLAB-safe variable name (≤ 63 chars)."""
    safe = re.sub(r"[^A-Za-z0-9_]", "_", name)
    if not safe or not safe[0].isalpha():
        safe = "ch_" + safe
    safe = safe[:60]
    n = seen.get(safe, 0)
    seen[safe] = n + 1
    return safe if n == 0 else f"{safe}_{n}"


def _delete(path: str) -> None:
    try:
        if os.path.isfile(path):
            os.remove(path)
    except OSError:
        pass


# --------------------------------------------------------------------------- #
# Split-output helpers
# --------------------------------------------------------------------------- #

def _get_time_range(mdf: Any) -> "tuple[float, float]":
    """Return (t_min, t_max) across all groups.  Returns (0.0, 0.0) when no data."""
    t_min, t_max = float("inf"), float("-inf")
    for i in range(len(mdf.groups)):
        try:
            ts = mdf.get_master(index=i)
            if ts is not None and len(ts) > 0:
                t_min = min(t_min, float(ts[0]))
                t_max = max(t_max, float(ts[-1]))
        except Exception:  # noqa: BLE001
            continue
    if t_min == float("inf"):
        return 0.0, 0.0
    return t_min, t_max


def _split_time_windows(
    t_min: float,
    t_max: float,
    period: float,
    first_offset: float,
) -> "list[tuple[float, float]]":
    """Generate non-overlapping (start, stop) windows covering [t_min, t_max].

    The first split boundary falls at ``t_min + first_offset``.
    When *first_offset* is 0, windows start at *t_min* with uniform spacing of
    *period*.  A positive *first_offset* produces an initial shorter window
    [t_min, t_min+first_offset) before the regular windows begin.
    """
    if period <= 0:
        return [(t_min, t_max)]

    windows: list[tuple[float, float]] = []
    first_boundary = t_min + first_offset

    # Optional partial first window before the first boundary
    if first_offset > 0 and first_boundary < t_max:
        windows.append((t_min, first_boundary))

    # Regular windows from first_boundary to t_max
    t = first_boundary if first_offset > 0 else t_min
    while t < t_max:
        windows.append((t, min(t + period, t_max)))
        t += period

    return windows if windows else [(t_min, t_max)]


def _estimate_bps(mdf: Any, source_path: str, duration: float) -> float:
    """Estimate source bytes-per-second to auto-calculate a size-based split period.

    Uses the actual source file size when available; falls back to a rough
    8-bytes-per-sample estimate when the path is missing or inaccessible.
    Returns 0 when estimation is impossible.
    """
    if duration <= 0:
        return 0.0
    if source_path and os.path.isfile(source_path):
        return os.path.getsize(source_path) / duration
    # Rough fallback: count cycles_nr × channels × 8 bytes
    total = sum(
        int(getattr(g.channel_group, "cycles_nr", 0) or 0) * len(g.channels)
        for g in mdf.groups
    )
    return (total * 8) / duration if total > 0 else 0.0


def _insert_suffix(output_path: str, suffix: str) -> str:
    """Insert *suffix* before the file extension.

    Example: _insert_suffix("/out/file.mat", "_T00060s") → "/out/file_T00060s.mat"
    """
    stem, ext = os.path.splitext(output_path)
    return f"{stem}{suffix}{ext}"


class _ChunkProxy:
    """Forwards cancel checks and _cleanup references to a parent _Job while
    suppressing per-group done/total updates from inner export functions.
    This lets each chunk's export function run without overwriting the chunk-level
    progress tracked by the parent job.
    """

    def __init__(self, parent: "_Job") -> None:
        self._parent  = parent
        # Share the parent's cleanup list so cancelled chunk files are purged.
        self._cleanup = parent._cleanup
        self.error:  "str | None" = None
        self.status: str          = "running"

    @property
    def cancel_requested(self) -> bool:
        return self._parent.cancel_requested

    # Suppress total/done updates from inner export functions; the split
    # dispatcher maintains the parent job's counters itself.
    @property
    def total(self) -> int:
        return self._parent.total

    @total.setter
    def total(self, _: int) -> None:
        pass

    @property
    def done(self) -> int:
        return self._parent.done

    @done.setter
    def done(self, _: int) -> None:
        pass

    def request_cancel(self) -> None:
        self._parent.request_cancel()


def _build_chunk_filter(
    chunk_mdf: Any,
    name_pair_filter: "set[tuple[str, str]] | None",
) -> "set[tuple[int, str]] | None":
    """Rebuild a (group_index, channel_name) filter_set for a per-chunk decoded MDF.

    After cut-then-decode each chunk's decoded MDF is independent; messages
    absent from the time window may not appear, so group indices differ from those
    in the full decoded MDF.  This function maps the saved (acq_name, channel_name)
    pairs back to the chunk's actual group indices.

    Returns ``None`` when no name-pair filter is active (export everything).
    Returns an empty set when none of the requested channels are present in the
    chunk (the chunk should produce an empty/skipped output).
    """
    if name_pair_filter is None:
        return None
    result: set[tuple[int, str]] = set()
    for i, group in enumerate(chunk_mdf.groups):
        cg  = group.channel_group
        acq = str(getattr(cg, "acq_name", "") or "")
        for ch in group.channels:
            ch_name = str(ch.name or "")
            if ch_name and (acq, ch_name) in name_pair_filter:
                result.add((i, ch_name))
    return result


def _absolute_suffix(t_mdf_s: float, recording_start: Any) -> str:
    """Return a ``_YYMMDD_HHMMSS`` filename suffix for the given MDF timestamp.

    *t_mdf_s* is the MDF-internal timestamp in seconds (as returned by
    ``get_master``).  *recording_start* is the ``datetime`` object stored in the
    MDF header (``mdf.header.start_time``); adding *t_mdf_s* to it gives the
    absolute point in time for this chunk's start.

    When *recording_start* is timezone-aware (asammdf typically stores it as UTC),
    the result is converted to the **local timezone** before formatting so that
    the filename suffix matches the wall-clock time the user sees in the UI.
    Naive datetimes are formatted as-is.
    Falls back to a zero-padded seconds offset when *recording_start* is
    unavailable or the conversion fails.
    """
    if recording_start is None:
        return f"_T{int(t_mdf_s):05d}s"
    try:
        from datetime import timedelta
        abs_dt = recording_start + timedelta(seconds=t_mdf_s)
        if getattr(abs_dt, "tzinfo", None) is not None:
            # Convert UTC (or any aware tz) → machine-local time, then drop tz.
            abs_dt = abs_dt.astimezone().replace(tzinfo=None)
        return abs_dt.strftime("_%y%m%d_%H%M%S")
    except Exception:  # noqa: BLE001
        return f"_T{int(t_mdf_s):05d}s"
